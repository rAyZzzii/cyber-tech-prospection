"""
Cyber-Tech Prospection — Générateur Excel structuré
Fichier source de vérité pour le suivi des offres de formation cybersécurité IDF

Usage:
    python3 scripts/generate_excel.py          # Crée/réinitialise le fichier
    python3 scripts/generate_excel.py --update  # Met à jour depuis data/offres.json
"""

import json
import os
import sys
from datetime import datetime, timedelta

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("openpyxl required. Install: pip install openpyxl")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
EXCEL_PATH = os.path.join(DATA_DIR, "cyber-tech_prospection.xlsx")
OFFRES_JSON = os.path.join(DATA_DIR, "offres.json")

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
URL_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FONT_RED = Font(color="9C0006")
FILL_ORANGE = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
FONT_ORANGE = Font(color="9C5700")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FONT_GREEN = Font(color="006100")

COLUMNS = [
    ("Titre du poste", 45),
    ("Établissement", 35),
    ("Type de mission", 20),
    ("Matières enseignées", 40),
    ("Niveau d'enseignement", 20),
    ("Localisation", 25),
    ("Date de publication", 18),
    ("Date limite", 18),
    ("URL de l'offre", 50),
    ("Email contact RH", 35),
    ("Statut", 22),
    ("Notes", 40),
]

STATUS_LIST = '"Non contacté,Email envoyé,Réponse positive,Refus,Sans réponse"'


def create_workbook():
    wb = Workbook()

    # ── Feuille 1 : Missions ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Missions"

    # Entêtes
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Ligne gelée
    ws.freeze_panes = "A2"

    # Autofiltre
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Validation de données colonne Statut (K)
    dv = DataValidation(type="list", formula1=STATUS_LIST, allow_blank=True)
    dv.error = "Choisis un statut valide"
    dv.errorTitle = "Statut invalide"
    dv.prompt = "Sélectionne le statut"
    dv.promptTitle = "Statut"
    ws.add_data_validation(dv)
    dv.add(f"K2:K5000")

    # Mise en forme conditionnelle
    # Rouge : Non contacté ET date > 7 jours (sur colonne A:K)
    ws.conditional_formatting.add(
        f"A2:L5000",
        CellIsRule(
            operator="equal",
            formula=['"Non contacté"'],
            fill=FILL_RED,
            font=FONT_RED,
        ),
    )
    ws.conditional_formatting.add(
        f"A2:L5000",
        CellIsRule(
            operator="equal",
            formula=['"Email envoyé"'],
            fill=FILL_ORANGE,
            font=FONT_ORANGE,
        ),
    )
    ws.conditional_formatting.add(
        f"A2:L5000",
        CellIsRule(
            operator="equal",
            formula=['"Réponse positive"'],
            fill=FILL_GREEN,
            font=FONT_GREEN,
        ),
    )

    # ── Feuille 2 : Tableau de bord ───────────────────────────────────────
    ws2 = wb.create_sheet("Tableau de bord")

    dashboard_data = [
        ("INDICATEUR", "VALEUR"),
        ("Total offres détectées", 0),
        ("", ""),
        ("Par source :", ""),
        ("  LinkedIn Jobs", 0),
        ("  Indeed France", 0),
        ("  Welcome to the Jungle", 0),
        ("  Lesformateurs.com", 0),
        ("  Malt", 0),
        ("  Sites carrières établissements", 0),
        ("", ""),
        ("Statuts :", ""),
        ("  Non contacté", 0),
        ("  Email envoyé", 0),
        ("  Réponse positive", 0),
        ("  Refus", 0),
        ("  Sans réponse", 0),
        ("", ""),
        ("Contacts envoyés", 0),
        ("Taux de réponse", "0%"),
        ("Opportunités en cours", 0),
        ("Dernière mise à jour", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    for row_idx, (label, value) in enumerate(dashboard_data, 1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)

        if row_idx == 1:
            cell_a.font = HEADER_FONT
            cell_a.fill = HEADER_FILL
            cell_b.font = HEADER_FONT
            cell_b.fill = HEADER_FILL
        else:
            cell_a.font = Font(name="Calibri", bold=True, size=10 if ":" in label else 11)
            cell_b.font = BODY_FONT
            cell_b.alignment = Alignment(horizontal="center")

        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 25

    wb.save(EXCEL_PATH)
    print(f"✅ Fichier Excel créé : {EXCEL_PATH}")
    return wb


def update_from_json():
    """Met à jour le fichier Excel depuis le JSON de veille."""
    if not os.path.exists(OFFRES_JSON):
        print("⚠️  Aucun fichier offres.json trouvé. Création Excel vierge.")
        create_workbook()
        return

    with open(OFFRES_JSON, "r", encoding="utf-8") as f:
        offres = json.load(f)

    if not os.path.exists(EXCEL_PATH):
        print("⚠️  Excel non trouvé. Création...")
        wb = create_workbook()
    else:
        wb = load_workbook(EXCEL_PATH)

    ws = wb["Missions"]

    # Nettoyer les lignes existantes (garder entête)
    ws.delete_rows(2, ws.max_row)

    for offre in offres:
        row = ws.max_row + 2 if ws.max_row >= 1 else 2  # après l'entête
        values = [
            offre.get("titre", ""),
            offre.get("etablissement", ""),
            offre.get("type_mission", ""),
            offre.get("matieres", ""),
            offre.get("niveau", ""),
            offre.get("localisation", ""),
            offre.get("date_publication", ""),
            offre.get("date_limite", ""),
            offre.get("url", ""),
            offre.get("email_contact", ""),
            offre.get("statut", "Non contacté"),
            offre.get("notes", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER
            if col_idx == 9 and val.startswith("http"):
                cell.font = URL_FONT

    # Mettre à jour le tableau de bord
    ws2 = wb["Tableau de bord"]
    total = len(offres)
    sources = {}
    statuses = {}
    for o in offres:
        src = o.get("source", "Inconnue")
        sources[src] = sources.get(src, 0) + 1
        st = o.get("statut", "Non contacté")
        statuses[st] = statuses.get(st, 0) + 1

    mapped = {
        "Total offres détectées": total,
        "  LinkedIn Jobs": sources.get("linkedin", 0),
        "  Indeed France": sources.get("indeed", 0),
        "  Welcome to the Jungle": sources.get("wttj", 0),
        "  Lesformateurs.com": sources.get("lesformateurs", 0),
        "  Malt": sources.get("malt", 0),
        "  Sites carrières établissements": sources.get("carrieres", 0),
        "  Non contacté": statuses.get("Non contacté", 0),
        "  Email envoyé": statuses.get("Email envoyé", 0),
        "  Réponse positive": statuses.get("Réponse positive", 0),
        "  Refus": statuses.get("Refus", 0),
        "  Sans réponse": statuses.get("Sans réponse", 0),
        "Contacts envoyés": statuses.get("Email envoyé", 0),
        "Taux de réponse": f"{round(statuses.get('Réponse positive', 0) / max(total, 1) * 100)}%"
            if total > 0 else "0%",
        "Opportunités en cours": statuses.get("Réponse positive", 0)
            + statuses.get("Email envoyé", 0),
        "Dernière mise à jour": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }

    for row_idx in range(2, ws2.max_row + 1):
        label = ws2.cell(row=row_idx, column=1).value
        if label and label.strip() in mapped:
            ws2.cell(row=row_idx, column=2, value=mapped[label.strip()])

    wb.save(EXCEL_PATH)
    print(f"✅ Excel mis à jour : {EXCEL_PATH} ({total} offres)")

    # Auto-export CSV
    import csv
    csv_path = os.path.join(DATA_DIR, "cyber-tech_prospection.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "Titre du poste", "Établissement", "Type de mission",
            "Matières enseignées", "Niveau d'enseignement", "Localisation",
            "Date de publication", "Date limite", "URL",
            "Email contact RH", "Statut", "Notes"
        ])
        for o in offres:
            w.writerow([
                o.get("titre", ""), o.get("etablissement", ""),
                o.get("type_mission", ""), o.get("matieres", ""),
                o.get("niveau", ""), o.get("localisation", ""),
                o.get("date_publication", ""), o.get("date_limite", ""),
                o.get("url", ""), o.get("email_contact", ""),
                o.get("statut", "Non contacté"), o.get("notes", "")
            ])
    print(f"✅ CSV mis à jour : {csv_path} ({total} offres)")


if __name__ == "__main__":
    if "--update" in sys.argv:
        update_from_json()
    else:
        create_workbook()